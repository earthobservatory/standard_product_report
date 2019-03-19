#!/usr/bin/env python

'''
Contains functions for writing Excel files for the Standard Product Report
'''
import re
import pickle
import hashlib
from openpyxl import Workbook
import dateutil.parser

def generate(aoi, acqs, slcs, acq_lists, ifg_cfgs, ifgs, audit_trail):
    '''ingests the various products and stages them by track for generating worksheets'''
    # unique tracks based on acquisition list
    unique_tracks = acq_lists.keys()
    for track in unique_tracks:
        print('generating workbook for track {}'.format(track))
        generate_track(track, aoi, acqs.get(track, []), slcs.get(track, []), acq_lists.get(track, []), ifg_cfgs.get(track, []), ifgs.get(track, []), audit_trail.get(track, []))

def generate_track(track, aoi, acqs, slcs, acq_lists, ifg_cfgs, ifgs, audit_trail):
    '''generates excel sheet for given track, inputs are lists'''
    # stage products
    filename = '{}_T{}.xlsx'.format(aoi.get('_id', 'AOI'), track)
    acq_dct = convert_to_dict(acqs) # converts to dict based on id
    slc_dct = convert_to_dict(slcs) # converts to dict based on id
    acq_map = map_acqs_to_slcs(acqs) # converts acquisition ids to slc ids
    slc_map = map_slcs_to_acqs(acqs) # converts slc ids to acq_ids
    acq_list_dct = convert_to_hash_dict(acq_lists, conversion_dict=acq_map) # converts dict where key is hash of master/slave slc ids
    ifg_cfg_dct = convert_to_hash_dict(ifg_cfgs, conversion_dict=acq_map) # converts dict where key is hash of master/slave slc ids
    ifg_dct = convert_to_hash_dict(ifgs) # converts dict where key is hash of master/slave slc ids

    # generate the acquisition sheet
    wb = Workbook()
    ws1 = wb.create_sheet("Enumerated Products")
    all_missing_slcs = [] # list of missing slcs by acquisition id
    titlerow = ['acquisition-list id', 'slcs localized?', 'ifg-cfg generated?', 'ifg generated?', 'missing slc ids', 'missing acq ids']
    ws1.append(titlerow)
    # for each acquisition list, determine relevant metrics
    for hkey in acq_list_dct.keys():
        obj = acq_list_dct.get(hkey)
        acqlistid = obj.get('_source', {}).get('id', 'No acquisition id found')
        slcs_are_localized = is_covered(obj, slc_dct) # True/False if SLCs are localized
        missing_acq_str = ''
        missing_slc_str = ''
        if not slcs_are_localized:
            missing_slcs = get_missing_slcs(obj, acq_map, slc_dct) # get list of any missing slc ids
            all_missing_slcs.extend(missing_slcs) # add to master list for later
            missing_slc_str = ' '.join(missing_slcs)
            missing_acqs = [slc_map.get(x, False) for x in missing_slcs]
            missing_acq_str = ' '.join(missing_acqs)
        row = [acqlistid, slcs_are_localized, in_dict(hkey, ifg_cfg_dct), in_dict(hkey, ifg_dct), missing_slc_str, missing_acq_str]
        ws1.append(row)
    # generate missing slc list
    ws2 = wb.create_sheet("Missing SLCs")
    all_missing_slcs = sorted(list(set(all_missing_slcs)))
    title_row = ['slc id', 'acquisition id', 'starttime', 'endtime']
    ws2.append(title_row)
    for slc_id in all_missing_slcs:
        acq_id = slc_map.get(slc_id)
        acq_obj = acq_dct.get(acq_id, {})
        starttime = acq_obj.get('_source', {}).get('starttime', '-')
        endtime = acq_obj.get('_source', {}).get('endtime', '-')
        row = [slc_id, acq_id, starttime, endtime]
        ws2.append(row)
    #determine all date pairs that should be generated
    ws3 = wb.create_sheet('Enumerated Date Pairs')
    all_date_pairs = []
    title_row = ['expected date pairs']
    ws3.append(title_row)
    for key in acq_list_dct.keys():
        acq_list = acq_list_dct[key]
        st = dateutil.parser.parse(acq_list.get('_source').get('starttime')).strftime('%Y%m%d')
        et = dateutil.parser.parse(acq_list.get('_source').get('endtime')).strftime('%Y%m%d')
        ts = '{}-{}'.format(et, st)
        all_date_pairs.append(ts)
    for dt in list(set(all_date_pairs)).sort():
        ws3.append([dt])
    #all acquisitions
    ws4 = wb.create_sheet('Acquisitions')
    title_row = ['acquisition id', 'starttime', 'endtime']
    for key in sorted(acq_dct.keys()):
        acq = acq_dct[key]
        acq_id = acq.get('_id', 'UNKNOWN')
        acq_st = acq.get('_source', {}).get('starttime', False)
        acq_et = acq.get('_source', {}).get('endttime', False)
        ws4.append([acq_id, acq_st, acq_et])
    #all slcs
    ws5 = wb.create_sheet('Localized SLCs')
    title_row = ['slc id', 'starttime', 'endtime']
    for key in sorted(slc_dct.keys()):
        slc = slc_dct[key]
        slc_id = slc.get('_id', 'UNKNOWN')
        slc_st = slc.get('_source', {}).get('starttime', False)
        slc_et = slc.get('_source', {}).get('endttime', False)
        ws5.append([slc_id, slc_st, slc_et])
    #all ifg_cfgs
    ws6 = wb.create_sheet('IFG CFGs')
    title_row = ['ifg-cfg id', 'starttime', 'endtime']
    for key in ifg_cfg_dct.keys():
        slc = ifg_cfg_dct[key]
        slc_id = slc.get('_id', 'UNKNOWN')
        slc_st = slc.get('_source', {}).get('starttime', False)
        slc_et = slc.get('_source', {}).get('endttime', False)
        ws6.append([slc_id, slc_st, slc_et])
    #all ifgs
    ws7 = wb.create_sheet('IFGs')
    title_row = ['ifg id', 'starttime', 'endtime']
    for key in ifg_dct.keys():
        slc = ifg_dct[key]
        slc_id = slc.get('_id', 'UNKNOWN')
        slc_st = slc.get('_source', {}).get('starttime', False)
        slc_et = slc.get('_source', {}).get('endttime', False)
        ws7.append([slc_id, slc_st, slc_et])
    #audit trail
    ws8 = wb.create_sheet('Audit Trail')
    #just write all keys
    title_row = audit_trail[0].get('_source', {}).get('metadata', {}).keys()
    for x in ['union_geojson', 'context']:
        title_row.remove(x)
    ws8.append(title_row)
    for element in audit_trail:
        met = element.get('_source', {}).get('metadata', {})
        publish_row = []
        for key in title_row:
            val = met.get(key, '')
            publish_row.append(val)
        ws8.append(publish_row) 
    wb.save(filename)

def in_dict(hsh, dct):
    '''returns true if the hash input is a key in the input dict'''
    rslt = dct.get(hsh, False)
    if rslt is False:
        return False
    return True

def convert_to_hash_dict(obj_list, conversion_dict=False):
    '''converts the list into a dict of objects where the keys are a hash of their master & slave slcs. if the entry
       is acquisitions, uses a conversion dict to convert to slc ids'''
    out_dict = {}
    for obj in obj_list:
        master = obj.get('_source', {}).get('metadata', {}).get('master_scenes', [])
        slave = obj.get('_source', {}).get('metadata', {}).get('slave_scenes', [])
        if conversion_dict:
            master = [conversion_dict.get(x, False) for x in master]
            slave = [conversion_dict.get(x, False) for x in slave]
        master = pickle.dumps(sorted(master))
        slave = pickle.dumps(sorted(slave))
        hsh = '{}_{}'.format(hashlib.md5(master).hexdigest(), hashlib.md5(slave).hexdigest())
        out_dict[hsh] = obj
    return out_dict

def is_covered(obj, slc_dct):
    '''returns True if the SLCs are in slc_dct, False otherwise'''
    master = obj.get('_source', {}).get('metadata', {}).get('master_scenes', [])
    slave = obj.get('_source', {}).get('metadata', {}).get('slave_scenes', [])
    slc_list = list(set(master).union(set(slave)))
    for slc_id in slc_list:
        if slc_dct.get(slc_id, False) is False:
            return False
    return True

def get_missing_slcs(obj, acq_map, slc_dct):
    '''returns the slc ids enumerated in the object that are not contained in the slc dict'''
    master = obj.get('_source', {}).get('metadata', {}).get('master_scenes', [])
    slave = obj.get('_source', {}).get('metadata', {}).get('slave_scenes', [])
    acq_ids = list(set(master).union(set(slave)))
    #convert the acquisition ids to slc ids
    slc_ids = [acq_map.get(x, False) for x in acq_ids]
    #if the slc ids are not in the slc dict
    missing = []
    for slc_id in slc_ids:
        if slc_dct.get(slc_id, False) is False:
            missing.append(slc_id)
    return missing

def convert_to_dict(input_list):
    '''attempts to convert the input list to a dict where the keys are object_id'''
    out_dict = {}
    for obj in input_list:
        obj_id = obj.get('_source', {}).get('id', False)
        out_dict[obj_id] = obj
    return out_dict

def convert_to_dt_dict(input_list):
    '''attempts to convert the input list to a dict where the keys are object_id'''
    out_dict = {}
    for obj in input_list:
        starttime = parse_start_time(obj)
        out_dict[starttime] = obj
    return out_dict
    
def parse_start_time(obj):
    '''gets start time'''
    st = obj.get('_source', {}).get('starttime', False)
    return dateutil.parser.parse(st).strftime('%Y-%m-%dT%H:%M:%S')

def parse_from_fn(obj_string):
    '''parses starttime from filename string'''
    reg = '([1-2][0-9]{7}T[0-9]{6})'
    dt = dateutil.parser.parse(re.findall(reg, obj_string)[0])
    return dt.strftime('%Y-%m-%dT%H:%M:%S')

def parse_slc_id(obj):
    '''returns the slc identifier for the acquisition list product'''
    obj_type = obj.get('_source', {}).get('dataset', False)
    if obj_type == 'acquisition-S1-IW_SLC':
        return obj.get('_source', {}).get('metadata', {}).get('dataset', False)
    if obj_type == 'S1-IW_SLC':
        return obj.get('_source', {}).get('id')
    return False

def map_slcs_to_acqs(acqs):
    '''returns a dict that takes in an SLC id and returns the associated acq id'''
    mapping_dict = {}
    for acq in acqs:
        slc_id = parse_slc_id(acq)
        acq_id = acq.get('_source', {}).get('id', False)
        mapping_dict[slc_id] = acq_id
    return mapping_dict
        
def map_acqs_to_slcs(acqs):
    '''returns a dict that takes in an acq id and returns the associated slc id'''
    mapping_dict = {}
    for acq in acqs:
        slc_id = parse_slc_id(acq)
        acq_id = acq.get('_source', {}).get('id', False)
        mapping_dict[acq_id] = slc_id
    return mapping_dict
